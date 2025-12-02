import hashlib
import json
import sys
import time
from pathlib import Path
from typing import Dict, List

import streamlit as st

from app.components.layout import init_page

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    load_workbook = None
    Workbook = None
    Font = None
    PatternFill = None
    Alignment = None

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

from app.components.autofill import d
from app.components.file_utils import build_step_filename, snapshot_to_pdf
from app.components.project_state import (
    ensure_dirs,
    ensure_local_path,
    load_manifest,
    project_subdir,
    register_chain_tx,
    register_file,
    resolve_stored_path,
    update_project_meta,
    use_project,
)
from app.components.web3_client import send_log_tx


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


COLUMN_HEADERS = [
    "Lfd.Nr.",
    "Parental cell line",
    "Cell line name",
    "Link Form Z",
    "IRIS ID",
    "Gene symbol (NCBI ID)",
    "crRNA/sgRNA sequence",
    "Donor/Vector (ssODN seq / Addgene cat. number)",
    "Modification Type",
    "Modification description",
    "Zygosity",
    "Date of Registration",
    "RG",
    "Disease",
]


def evaluate_readiness(rows: List[Dict[str, str]]) -> Dict[str, List[str] | bool]:
    issues: List[str] = []
    warnings: List[str] = []
    required = [
        "Lfd.Nr.",
        "Parental cell line",
        "Cell line name",
        "IRIS ID",
        "Gene symbol (NCBI ID)",
        "crRNA/sgRNA sequence",
        "Modification Type",
        "Modification description",
        "Date of Registration",
    ]
    for idx, row in enumerate(rows, start=1):
        for key in required:
            if not (row.get(key) or "").strip():
                issues.append(f"Row {idx}: '{key}' is required.")
        if not (row.get("Donor/Vector (ssODN seq / Addgene cat. number)") or "").strip():
            warnings.append(f"Row {idx}: Donor/Vector information missing.")
        if not (row.get("RG") or "").strip():
            warnings.append(f"Row {idx}: Risk group (RG) not specified.")
    return {"ready": not issues, "issues": issues, "warnings": warnings}

init_page("Step 10 - LAGESO Compliance Manager")
st.title("Step 10 - LAGESO Compliance Manager")
st.caption("Edit compliance entries, save snapshots, and generate the official Excel workbook.")

project_id = use_project("Project")
if not project_id:
    st.warning("Select a project from the sidebar to continue.")
    st.stop()

manifest = load_manifest(project_id)
meta = manifest.get("meta", {})
compliance = meta.get("compliance", {})

project_root = ensure_dirs(project_id)
snapshots_root = project_root / "snapshots"
reports_dir = project_subdir(project_id, "reports", "lageso")
snapshot_dir = project_subdir(project_id, "snapshots", "lageso")

templates_dir = ROOT / "data" / "templates"
available_templates = sorted(templates_dir.glob("*.xlsm"))
saved_template = Path(compliance.get("excel_template_path", "")) if compliance.get("excel_template_path") else None
if available_templates:
    template_labels = [str(p) for p in available_templates]
    default_index = 0
    if saved_template and str(saved_template) in template_labels:
        default_index = template_labels.index(str(saved_template))
    template_choice = st.selectbox("Excel template (.xlsm)", template_labels, index=default_index)
else:
    template_choice = st.text_input(
        "Excel template (.xlsm)",
        str(saved_template or (templates_dir / "13_15 SD Proj 1-3.xlsm")),
    )

template_path = Path(template_choice)
method = st.selectbox("Sheet / method", ["TALENs", "CRISPR plasmids", "CRISPR RNPs"], index=2)
row_defaults = compliance.get("excel_rows", {})
start_row = st.number_input(
    "Start row to write (1-indexed)",
    min_value=1,
    value=row_defaults.get(method, {}).get("start_row", 5),
    step=1,
)


def latest_json(folder: Path) -> Dict:
    if not folder.exists():
        return {}
    try:
        files = sorted(folder.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    except Exception:
        return {}
    for file in files:
        try:
            if not file.exists():
                continue
            return json.loads(file.read_text(encoding="utf-8"))
        except (FileNotFoundError, OSError, PermissionError, json.JSONDecodeError):
            continue
        except Exception:
            continue
    return {}


# Check if force refresh is requested
force_refresh = st.session_state.get("lageso_force_refresh", False)
if force_refresh:
    st.session_state.lageso_force_refresh = False

# Always start fresh - clear saved rows unless user explicitly saved this session
if "lageso_user_saved" not in st.session_state:
    row_defaults = {}

# Load data from all project steps
charter = latest_json(snapshots_root / "charter")
design = latest_json(snapshots_root / "design")
delivery = latest_json(snapshots_root / "delivery")
assessment = latest_json(snapshots_root / "assessment")
cloning = latest_json(snapshots_root / "cloning")
screening = latest_json(snapshots_root / "screening")
master_bank = latest_json(snapshots_root / "master_bank_registry")

# Extract charter information
primary_cell_line = charter.get("cell_line", meta.get("cell_line", ""))
edit_program = charter.get("edit_program", meta.get("edit_program", ""))
compliance_scope = charter.get("compliance_scope", meta.get("compliance_scope", ""))
disease = charter.get("disease_relevance", meta.get("disease_relevance", meta.get("disease", "")))
owner = charter.get("owner", meta.get("owner", ""))

# Extract design information
design_mutation = design.get("mutation") or {}
gene_symbol = design_mutation.get("gene", meta.get("gene", ""))
edit_intent = design_mutation.get("edit_intent", "")

# Extract gRNAs (up to 2 from selected_guides and additional_guides)
design_guides = design.get("selected_guides") or []
additional_guides = design.get("additional_guides") or []
all_guides = design_guides + additional_guides
grna1 = all_guides[0].get("sequence", "") if len(all_guides) > 0 else ""
grna1_id = all_guides[0].get("id", "") if len(all_guides) > 0 else ""
grna2 = all_guides[1].get("sequence", "") if len(all_guides) > 1 else ""
grna2_id = all_guides[1].get("id", "") if len(all_guides) > 1 else ""

# Format: ID: sequence
sgrna1_formatted = f"{grna1_id}: {grna1}" if grna1_id and grna1 else grna1
sgrna2_formatted = f"{grna2_id}: {grna2}" if grna2_id and grna2 else grna2
sgrna_sequences = f"{sgrna1_formatted}; {sgrna2_formatted}" if sgrna2_formatted else sgrna1_formatted

# Extract donors from 'donors' array (up to 2)
design_donors = design.get("donors") or []
donor1 = design_donors[0] if len(design_donors) > 0 else {}
donor2 = design_donors[1] if len(design_donors) > 1 else {}

donor1_sequence = donor1.get("sequence", "")
donor1_type = donor1.get("donor_type", "")
donor2_sequence = donor2.get("sequence", "")
donor2_type = donor2.get("donor_type", "")

# Also check old donor field for backward compatibility
design_primers = design.get("primer_pairs") or []
design_donor = design.get("donor") or {}
if not donor1_sequence:
    donor1_sequence = design_donor.get("sequence", "")
    donor1_type = design_donor.get("donor_type", design_donor.get("type", ""))

vector_name = design.get("vector_name", design_donor.get("name", ""))
addgene_number = design.get("addgene_number", design_donor.get("addgene_cat", ""))

# Extract master bank data - get all cell lines registered
mb_entries = master_bank.get("entries") or []
cell_line_names = []
iris_ids = []
registration_dates = []
zygosities = []

for entry in mb_entries:
    mb_data = entry.get("data", {})
    cell_name = mb_data.get("HPSCReg_Name") or mb_data.get("IRIS_ID", "")
    if cell_name:
        cell_line_names.append(cell_name)
    iris_id = mb_data.get("IRIS_ID", "")
    if iris_id:
        iris_ids.append(iris_id)
    reg_date = mb_data.get("Date_Registered") or mb_data.get("date_registered", "")
    if reg_date:
        registration_dates.append(str(reg_date))
    # Use verification field for Zygosity
    zyg = mb_data.get("verification", mb_data.get("Zygosity", ""))
    if zyg:
        zygosities.append(zyg)

# Get first entry for defaults
mb_data = mb_entries[0].get("data", {}) if mb_entries else {}
first_cell_line = cell_line_names[0] if cell_line_names else ""
first_iris_id = iris_ids[0] if iris_ids else project_id
first_date = registration_dates[0] if registration_dates else time.strftime("%Y-%m-%d")
first_zygosity = zygosities[0] if zygosities else ""

# Use full project ID (e.g., "59753 (61377)")
bank_identifier = project_id

cell_line_name = first_cell_line or meta.get("cell_line") or compliance.get("cell_line", "")
operator_name = (
    compliance.get("responsible_person") or owner or meta.get("owner") or meta.get("contacts", {}).get("operator", "")
)

values_preview = {
    "ProjectID": project_id,
    "CellLine": cell_line_name,
    "Gene": meta.get("gene") or design.get("mutation", {}).get("gene", ""),
    "Edit": design.get("mutation", {}).get("edit_intent", ""),
    "Operator": operator_name,
    "Date": time.strftime("%Y-%m-%d"),
    "gRNA1": d(design_guides, 0, "sequence", default=""),
    "gRNA2": d(design_guides, 1, "sequence", default=""),
    "PrimerFw": d(design_primers, 0, "forward", default=""),
    "PrimerRv": d(design_primers, 0, "reverse", default=""),
    "Amplicon_bp": d(design_primers, 0, "expected_amplicon_bp", default=""),
    "DonorType": design_donor.get("donor_type") or design_donor.get("type", ""),
    "DonorSeq": design_donor.get("sequence", ""),
    "AssayTool": assessment.get("tool_used") or "",
    "TotalIndelPct": assessment.get("total_indel_pct")
    or d(screening, "clones", 0, "sanger", "total_indel_pct", default=""),
    "Positives": ", ".join(screening.get("positives", []))
    if isinstance(screening.get("positives"), list)
    else "",
    "BankID": bank_identifier,
    "Passage": mb_data.get("Bank_size") or compliance.get("passage", ""),
}

st.markdown("### Current values derived from snapshots")
st.json(values_preview)

# Build comprehensive auto-filled default row
# Format donor field with both donors if available
# Create pattern: "GeneName_EditType_ssODN1: sequence1; GeneName_EditType_ssODN2: sequence2"
donor_field = ""

# Extract gene name from gene_symbol (remove NCBI ID if present)
gene_name_short = gene_symbol.split("(")[0].strip() if gene_symbol else "Gene"
gene_name_short = gene_name_short.split()[0]  # Get first word only

# Create donor prefix based on gene and edit type
edit_type_short = edit_intent.replace("-", "_") if edit_intent else "Edit"

def format_donor_with_id(sequence: str, donor_num: int, gene: str, edit: str) -> str:
    """Format donor sequence with ID prefix"""
    if not sequence:
        return ""
    # Check if sequence already has a prefix pattern (contains colon in first 50 chars)
    if ":" in sequence[:50]:
        return sequence  # Already has prefix
    # Add prefix: GeneName_EditType_ssODN#: sequence
    return f"{gene}_{edit}_ssODN{donor_num}: {sequence}"

if donor1_type == "ssODN" and donor1_sequence:
    donor1_formatted = format_donor_with_id(donor1_sequence, 1, gene_name_short, edit_type_short)
    donor_field = donor1_formatted
    
    if donor2_type == "ssODN" and donor2_sequence:
        donor2_formatted = format_donor_with_id(donor2_sequence, 2, gene_name_short, edit_type_short)
        donor_field = f"{donor1_formatted}; {donor2_formatted}"
elif vector_name or addgene_number:
    donor_field = f"{vector_name} (Addgene #{addgene_number})" if addgene_number and vector_name else vector_name
elif donor1_sequence:
    donor_field = donor1_sequence

default_row = {
    "Lfd.Nr.": compliance.get("sequential_number", "1"),
    "Parental cell line": primary_cell_line,
    "Cell line name": cell_line_name,
    "Link Form Z": compliance.get("link_form_z", ""),
    "IRIS ID": bank_identifier,
    "Gene symbol (NCBI ID)": gene_symbol,
    "crRNA/sgRNA sequence": sgrna_sequences,
    "Donor/Vector (ssODN seq / Addgene cat. number)": donor_field,
    "Modification Type": edit_intent or edit_program,
    "Modification description": compliance.get("modification_description", ""),
    "Zygosity": first_zygosity,
    "Date of Registration": first_date,
    "RG": compliance.get("risk_classification", "1"),
    "Disease": disease,
}

# Create multiple rows - one per Master Bank entry
default_rows = []
if mb_entries:
    for idx, entry in enumerate(mb_entries, start=1):
        mb_data_entry = entry.get("data", {})
        entry_cell_line = mb_data_entry.get("HPSCReg_Name") or mb_data_entry.get("IRIS_ID", "")
        entry_iris_id = mb_data_entry.get("IRIS_ID", "")
        entry_date = mb_data_entry.get("Date_Registered") or mb_data_entry.get("date_registered", "")
        # Use verification field for Zygosity
        entry_zygosity = mb_data_entry.get("verification", mb_data_entry.get("Zygosity", ""))
        
        # Use full project ID for IRIS ID
        entry_iris_full = project_id
        
        default_rows.append({
            "Lfd.Nr.": compliance.get("sequential_number", str(idx)),
            "Parental cell line": primary_cell_line,
            "Cell line name": entry_cell_line,
            "Link Form Z": compliance.get("link_form_z", ""),
            "IRIS ID": entry_iris_full,
            "Gene symbol (NCBI ID)": gene_symbol,
            "crRNA/sgRNA sequence": sgrna_sequences,
            "Donor/Vector (ssODN seq / Addgene cat. number)": donor_field,
            "Modification Type": edit_intent or edit_program,
            "Modification description": compliance.get("modification_description", ""),
            "Zygosity": entry_zygosity,
            "Date of Registration": str(entry_date) if entry_date else time.strftime("%Y-%m-%d"),
            "RG": compliance.get("risk_classification", "1"),
            "Disease": disease,
        })
else:
    # No master bank entries, use default single row
    default_rows = [default_row]

# Show auto-fill sources
with st.expander("ℹ️ Auto-filled data sources", expanded=False):
    st.markdown("""
    **Data extracted from:**
    - **Charter**: Parental cell line, Gene symbol, NCBI ID, Modification type/description, Disease, Risk group
    - **Design Log**: sgRNA sequences, Donor/Vector info, Addgene numbers
    - **Master Bank Registry**: Cell line names, IRIS IDs, Registration dates, Zygosity
    """)
    if charter:
        st.markdown(f"✓ Charter data loaded")
    if design:
        st.markdown(f"✓ Design data loaded")
    if master_bank:
        st.markdown(f"✓ Master Bank data loaded ({len(mb_entries)} entries)")

existing_rows = row_defaults.get(method, {}).get("rows") or []

# Always use fresh auto-filled data from snapshots
if not existing_rows:
    existing_rows = default_rows

# Snapshot availability diagnostics for user clarity
def _count_json(folder: Path) -> int:
    try:
        if not folder.exists():
            return 0
        return len(list(folder.glob("*.json")))
    except Exception:
        return 0

st.markdown("**Snapshot availability:**")
availability = [
    ("Charter", snapshots_root / "charter"),
    ("Design", snapshots_root / "design"),
    ("Delivery", snapshots_root / "delivery"),
    ("Assessment", snapshots_root / "assessment"),
    ("Cloning", snapshots_root / "cloning"),
    ("Screening", snapshots_root / "screening"),
    ("Master Bank", snapshots_root / "master_bank_registry"),
]
for label, folder in availability:
    st.write(f"{label}: {_count_json(folder)} snapshot(s)")

st.markdown("### Compliance table")

col1, col2 = st.columns([3, 1])
with col2:
    if st.button("🔄 Refresh from latest data"):
        # Set flag to force refresh on next rerun
        st.session_state.lageso_force_refresh = True
        # Clear old cached data
        st.session_state.pop(f"lageso_table_{method}", None)
        st.success("✓ Refreshing with latest project data...")
        st.rerun()

table_editor = st.data_editor(
    existing_rows,
    column_config={col: st.column_config.TextColumn() for col in COLUMN_HEADERS},
    num_rows="dynamic",
    use_container_width=True,
    key=f"lageso_table_{method}",
)
table_records = table_editor.to_dict("records") if hasattr(table_editor, "to_dict") else list(table_editor)

readiness = evaluate_readiness(table_records or [default_row])
st.markdown("Readiness status:")
st.write(readiness)


def save_compliance_snapshot(rows: List[Dict[str, str]]):
    timestamp = int(time.time())
    payload = {
        "project_id": project_id,
        "timestamp_unix": timestamp,
        "method": method,
        "start_row": start_row,
        "rows": rows or [default_row],
    }
    payload_bytes = json.dumps(payload, indent=2).encode("utf-8")
    digest = sha256_bytes(payload_bytes)
    outfile = ensure_local_path(snapshot_dir / f"{project_id}_lageso_compliance_{timestamp}_{digest[:12]}.json")
    outfile.write_bytes(payload_bytes)

    st.session_state["lageso_compliance_snapshot"] = {
        "digest": digest,
        "outfile": str(outfile),
        "metadata_uri": outfile.resolve().as_uri(),
        "payload": payload,
    }
    st.session_state["lageso_compliance_readiness"] = readiness

    register_file(
        project_id,
        "snapshots",
        {
            "step": "lageso_compliance",
            "path": str(outfile),
            "digest": digest,
            "timestamp": timestamp,
        },
    )
    pdf_filename = build_step_filename(project_id, "lageso-compliance", timestamp)
    pdf_path = ensure_local_path(reports_dir / pdf_filename)
    snapshot_to_pdf(payload, pdf_path, "LAGESO Compliance Snapshot")
    register_file(
        project_id,
        "reports",
        {
            "step": "lageso_compliance",
            "path": str(pdf_path),
            "timestamp": timestamp,
            "digest": sha256_bytes(pdf_path.read_bytes()),
            "type": "pdf",
        },
    )
    st.success("Snapshot saved and PDF generated.")
    with pdf_path.open("rb") as handle:
        st.download_button("Download compliance snapshot (PDF)", handle, file_name=pdf_path.name)

    update_project_meta(
        project_id,
        {
            "compliance": {
                "excel_template_path": str(template_path),
                "last_excel_method": method,
                "excel_rows": {
                    **row_defaults,
                    method: {
                        "start_row": start_row,
                        "rows": rows,
                    },
                },
                "sequential_number": rows[0].get("Lfd.Nr.", ""),
                "parental_cell_line": rows[0].get("Parental cell line", ""),
                "cell_line": rows[0].get("Cell line name", ""),
                "link_form_z": rows[0].get("Link Form Z", ""),
                "bank_id": rows[0].get("IRIS ID", ""),
                "gene_symbol_ncbi": rows[0].get("Gene symbol (NCBI ID)", ""),
                "donor_vector": rows[0].get("Donor/Vector (ssODN seq / Addgene cat. number)", ""),
                "modification_description": rows[0].get("Modification description", ""),
                "zygosity": rows[0].get("Zygosity", ""),
                "registration_date": rows[0].get("Date of Registration", ""),
                "risk_classification": rows[0].get("RG", ""),
                "disease": rows[0].get("Disease", ""),
            }
        },
    )


def generate_excel(rows: List[Dict[str, str]]):
    """Generate a standalone Excel file with just the compliance table"""
    if Workbook is None or load_workbook is None:
        st.error("openpyxl is required (`pip install openpyxl`).")
        return
    
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "LAGESO Compliance"
        
        # Write headers with styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(COLUMN_HEADERS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_name, "")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_length = len(str(header))
            for row_idx in range(2, len(rows) + 2):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, min(len(cell_value), 50))
            ws.column_dimensions[col_letter].width = max_length + 2
        
        # Save the workbook
        timestamp = int(time.time())
        out_path = reports_dir / f"LAGESO_Compliance_{project_id}_{timestamp}.xlsx"
        wb.save(out_path)
        
        register_file(
            project_id,
            "reports",
            {
                "step": "lageso_compliance",
                "path": str(out_path),
                "timestamp": timestamp,
                "label": "LAGESO Compliance Table",
                "type": "xlsx",
            },
        )
        
        st.success(f"✓ Generated compliance Excel with {len(rows)} row(s): {out_path.name}")
        with out_path.open("rb") as handle:
            st.download_button("📥 Download Compliance Excel", handle, file_name=out_path.name, key="compliance_excel_download")
    
    except Exception as e:
        st.error(f"Error generating Excel: {str(e)}")


def append_to_sd_template(rows: List[Dict[str, str]]):
    """Append LAGESO Compliance data to SD template Sheet 3 (CRISPR)"""
    if load_workbook is None:
        st.error("openpyxl is required (`pip install openpyxl`).")
        return
    
    sd_template_path = templates_dir / "13_15 SD Proj 1-3.xlsm"
    if not sd_template_path.exists():
        st.error(f"SD Template not found: {sd_template_path}")
        return

    try:
        from openpyxl.styles import Font, Alignment, Border, Side
        from copy import copy
        
        wb = load_workbook(sd_template_path, keep_vba=True)
        
        # Sheet 3 is named "CRISPR" (index 2, 0-based)
        if len(wb.sheetnames) < 3:
            st.error("SD template doesn't have 3 sheets")
            return
        
        ws = wb.worksheets[2]  # Third sheet (CRISPR)
        
        # Find the last row with data to append at the bottom
        last_row = 1  # Start from row 1
        for row in ws.iter_rows(min_row=1, max_col=1):  # Check column A
            if row[0].value is not None:
                last_row = row[0].row
        
        start_row = last_row + 1  # Append after the last row with data
        
        # Find a reference row with good formatting (row 4 is first data row)
        reference_row = 4 if last_row >= 4 else last_row
        
        # Column mapping for SD template Sheet 3
        sd_columns = {
            "Lfd.Nr.": 1,  # Column A
            "Parental cell line": 2,  # Column B
            "Cell line name": 3,  # Column C
            "Link Form Z": 4,  # Column D
            "IRIS ID": 5,  # Column E
            "Gene symbol (NCBI ID)": 6,  # Column F
            "crRNA/sgRNA sequence": 7,  # Column G
            "Donor/Vector (ssODN seq / Addgene cat. number)": 8,  # Column H
            "Modification Type": 9,  # Column I
            "Modification description": 10,  # Column J
            "Zygosity": 11,  # Column K
            "Date of Registration": 12,  # Column L
            "RG": 13,  # Column M
            "Disease": 14,  # Column N
        }
        
        # Write each row to the sheet with formatting
        for row_idx, row in enumerate(rows):
            excel_row = start_row + row_idx
            
            # Calculate sequential number for column A (Lfd.Nr.)
            # The number should be the row number minus the header rows (usually row 4 is the first data row, so it's number 1)
            sequential_number = excel_row - 3  # Row 4 becomes 1, row 5 becomes 2, etc.
            
            for col_name, col_num in sd_columns.items():
                try:
                    cell = ws.cell(row=excel_row, column=col_num)
                    # Check if this cell is part of a merged range
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            break
                    
                    if not is_merged:
                        # Set the value - use sequential number for Lfd.Nr. column
                        if col_name == "Lfd.Nr.":
                            cell.value = sequential_number
                        else:
                            cell.value = row.get(col_name, "")
                        
                        # Copy formatting from reference row
                        ref_cell = ws.cell(row=reference_row, column=col_num)
                        if ref_cell.font:
                            cell.font = copy(ref_cell.font)
                        if ref_cell.alignment:
                            cell.alignment = copy(ref_cell.alignment)
                        if ref_cell.border:
                            cell.border = copy(ref_cell.border)
                        if ref_cell.fill:
                            cell.fill = copy(ref_cell.fill)
                except (AttributeError, TypeError):
                    # Skip cells that can't be written to
                    continue
        
        # Save directly to the template file (overwrite/update it)
        wb.save(sd_template_path)
        
        timestamp = int(time.time())
        register_file(
            project_id,
            "reports",
            {
                "step": "lageso_sd_export",
                "path": str(sd_template_path),
                "timestamp": timestamp,
                "label": "SD Project 3 Template Updated",
                "type": "xlsm",
            },
        )
        
        st.success(f"✓ Appended {len(rows)} row(s) to SD template in templates folder")
        with sd_template_path.open("rb") as handle:
            st.download_button("📥 Download Updated SD Template", handle, file_name=sd_template_path.name, key="sd_download")
    
    except Exception as e:
        st.error(f"Error exporting to SD template: {str(e)}")


col_save, col_excel, col_sd, col_anchor = st.columns(4)

with col_save:
    if st.button("Save compliance snapshot"):
        save_compliance_snapshot(table_records)
        st.session_state.lageso_user_saved = True

with col_excel:
    if st.button("Generate Excel"):
        generate_excel(table_records)

with col_sd:
    if st.button("📤 Export to SD Proj 3"):
        append_to_sd_template(table_records)

with col_anchor:
    snapshot_state = st.session_state.get("lageso_compliance_snapshot")
    readiness_state = st.session_state.get("lageso_compliance_readiness", readiness)
    if snapshot_state:
        st.code(f"Hash: {snapshot_state['digest']}", language="text")
        if readiness_state["ready"]:
            st.success("Ready to anchor.")
        else:
            st.error("Resolve issues before anchoring:")
            for issue in readiness_state["issues"]:
                st.write(f"- {issue}")
            if readiness_state["warnings"]:
                st.caption("Warnings:")
                for warning in readiness_state["warnings"]:
                    st.caption(f"Warning: {warning}")
        disabled = not readiness_state["ready"]
        if st.button("Anchor snapshot", disabled=disabled):
            try:
                result = send_log_tx(
                    hex_digest=snapshot_state["digest"],
                    step="LAGESO Compliance",
                    metadata_uri=snapshot_state["metadata_uri"],
                )
            except Exception as exc:
                st.error(f"Anchoring failed: {exc}")
            else:
                st.success("Anchored on-chain.")
                st.write(f"Tx: {result['tx_hash']}")
                st.json(result["receipt"])
                register_chain_tx(
                    project_id,
                    {
                        "step": "lageso_compliance",
                        "tx_hash": result["tx_hash"],
                        "digest": snapshot_state["digest"],
                        "timestamp": int(time.time()),
                        "metadata_uri": snapshot_state["metadata_uri"],
                    },
                )
    else:
        st.info("Save a compliance snapshot to enable anchoring.")
